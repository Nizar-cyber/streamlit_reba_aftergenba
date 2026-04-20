"""
REBA Ergonomics Analyzer — Web App (Streamlit)
Converted from desktop (customtkinter) version.
Fungsi kamera realtime dihapus; hanya upload foto dari galeri.
"""

import cv2
import mediapipe as mp
import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime
from io import BytesIO
from PIL import Image

# ─── Konfigurasi halaman ─────────────────────────────────────────────────────
st.set_page_config(
    page_title="REBA Ergonomics Analyzer",
    page_icon="🦴",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CSS kustom ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
  /* Body & background */
  html, body, [data-testid="stAppViewContainer"] {
      background-color: #080812 !important;
      color: #CCCCDD;
  }
  [data-testid="stSidebar"] {
      background-color: #10101C !important;
  }
  /* Judul sidebar */
  .reba-title { font-size:2rem; font-weight:800; color:#7EB8F7; }
  .reba-subtitle { font-size:.75rem; color:#555577; margin-top:-10px; }
  /* Kartu skor */
  .score-card {
      border-radius:12px; padding:16px 20px; margin-bottom:8px; text-align:center;
  }
  /* Tabel hasil */
  .result-row { display:flex; justify-content:space-between; padding:4px 0;
                border-bottom:1px solid #2A2A3A; font-size:.9rem; }
  .result-label { color:#AAAAAA; }
  .result-value { font-weight:700; color:#EEEEFF; }
  /* Sembunyikan footer Streamlit */
  footer { visibility:hidden; }
</style>
""", unsafe_allow_html=True)

# =============================================
# MEDIAPIPE SETUP
# =============================================
@st.cache_resource
def load_pose():
    mp_pose = mp.solutions.pose
    return mp_pose, mp_pose.Pose(static_image_mode=True, min_detection_confidence=0.5)

mp_pose, pose_detector = load_pose()

# =============================================
# REBA LOOKUP TABLES (Hignett & McAtamney, 2000)
# =============================================
TABLE_A = [
    [[1,2,3,4],[1,2,3,4],[3,3,5,6]],
    [[2,3,4,5],[3,4,5,6],[4,5,6,7]],
    [[2,4,5,6],[4,5,6,7],[5,6,7,8]],
    [[3,5,6,7],[5,6,7,8],[6,7,8,9]],
    [[4,6,7,8],[6,7,8,9],[7,8,9,9]],
]
TABLE_B = [
    [[1,2,2],[1,2,3]],
    [[1,2,3],[2,3,4]],
    [[3,4,5],[4,5,5]],
    [[4,5,5],[5,6,7]],
    [[6,7,8],[7,8,8]],
    [[7,8,8],[8,9,9]],
]
TABLE_C = [
    [ 1, 1, 1, 2, 3, 3, 4, 5, 6, 7, 7, 7],
    [ 1, 2, 2, 3, 4, 4, 5, 6, 6, 7, 7, 8],
    [ 2, 3, 3, 3, 4, 5, 6, 7, 7, 8, 8, 8],
    [ 3, 4, 4, 4, 5, 6, 7, 8, 8, 9, 9, 9],
    [ 4, 4, 4, 5, 6, 7, 8, 8, 9, 9,10,10],
    [ 6, 6, 6, 7, 8, 8, 9, 9,10,10,11,11],
    [ 7, 7, 7, 8, 9, 9,10,10,11,11,11,12],
    [ 8, 8, 8, 9,10,10,11,11,12,12,13,13],
    [ 9, 9, 9,10,10,11,11,12,13,13,13,13],
    [10,10,10,11,11,12,12,13,13,14,14,14],
    [11,11,11,11,12,12,13,13,14,14,15,15],
    [12,12,12,12,12,13,13,14,14,15,15,15],
]

# =============================================
# HELPER WARNA
# =============================================
def risk_color_bgr(s):
    if s <= 1:    return (39,174,39)
    elif s <= 3:  return (50,205,50)
    elif s <= 7:  return (0,165,255)
    elif s <= 10: return (0,100,230)
    else:         return (0,50,220)

def seg_color(s):
    if s <= 1:   return (50,200,50)
    elif s == 2: return (0,200,200)
    elif s == 3: return (0,165,255)
    else:        return (0,80,220)

def risk_cat(s):
    if s == 1:    return ("Dapat Diabaikan", "#27AE60", "Tidak perlu tindakan")
    elif s <= 3:  return ("Rendah",          "#2ECC71", "Perubahan mungkin diperlukan")
    elif s <= 7:  return ("Sedang",          "#F39C12", "Investigasi lebih lanjut, perubahan segera")
    elif s <= 10: return ("Tinggi",          "#E67E22", "Investigasi dan implementasi perubahan segera")
    else:         return ("Sangat Tinggi",   "#E74C3C", "Implementasi perubahan SEGERA!")

# =============================================
# KALKULASI SUDUT
# =============================================
def calc_angle(a, b, c):
    a = np.array(a, float); b = np.array(b, float); c = np.array(c, float)
    ba = a - b; bc = c - b
    cos = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)

def trunk_flexion(shoulder, hip):
    v = np.array([hip[0]-shoulder[0], hip[1]-shoulder[1]])
    cos = np.dot(v, [0, 1]) / (np.linalg.norm(v) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)

def neck_flexion(ear, shoulder, hip):
    nv = np.array([ear[0]-shoulder[0], ear[1]-shoulder[1]])
    tv = np.array([shoulder[0]-hip[0], shoulder[1]-hip[1]])
    cos = np.dot(nv, tv) / (np.linalg.norm(nv)*np.linalg.norm(tv) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)

def upper_arm_angle(shoulder, elbow, hip):
    av = np.array([elbow[0]-shoulder[0], elbow[1]-shoulder[1]])
    tv = np.array([hip[0]-shoulder[0],   hip[1]-shoulder[1]])
    cos = np.dot(av, tv) / (np.linalg.norm(av)*np.linalg.norm(tv) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)

# =============================================
# REBA SCORING
# =============================================
def score_neck(a):  return 1 if a <= 20 else 2
def score_trunk(a):
    if a < 5:     return 1
    elif a <= 20: return 2
    elif a <= 60: return 3
    else:         return 4
def score_legs(ka):
    b = 1; f = 180 - ka
    if 30 <= f <= 60: b += 1
    elif f > 60:      b += 2
    return min(b, 4)
def score_ua(a):
    if a <= 20:   return 1
    elif a <= 45: return 2
    elif a <= 90: return 3
    else:         return 4
def score_la(a):
    f = 180 - a
    return 1 if 60 <= f <= 100 else 2
def score_wrist(d): return 1 if d <= 15 else 2

def tbl_a(t, n, l): return TABLE_A[max(1,min(5,t))-1][max(1,min(3,n))-1][max(1,min(4,l))-1]
def tbl_b(u, l, w): return TABLE_B[max(1,min(6,u))-1][max(1,min(2,l))-1][max(1,min(3,w))-1]
def tbl_c(a, b):    return TABLE_C[max(1,min(12,a))-1][max(1,min(12,b))-1]

def force_score(kg):
    try: v = float(kg)
    except: v = 0
    return 0 if v < 5 else (1 if v <= 10 else 2)

# =============================================
# GAMBAR SKELETON
# =============================================
def get_scale_params(w, h):
    ref = min(w, h); base = 600.0; s = ref / base
    return dict(
        bone  = max(1, int(2*s)),
        joint = max(3, int(4*s)),
        arc   = max(12, int(18*s)),
        font  = max(0.28, min(0.45, 0.35*s)),
        offset= max(35, int(45*s)),
        head  = 0.7,
        banner= max(0.38, min(0.55, 0.45*s)),
        legend= max(0.28, min(0.38, 0.32*s)),
    )

def draw_label(img, text, pos, color, scale=0.35):
    x, y = int(pos[0]), int(pos[1])
    font = cv2.FONT_HERSHEY_SIMPLEX
    thick = 1
    (tw, th), bl = cv2.getTextSize(text, font, scale, thick)
    p = 2
    cv2.rectangle(img, (x-p, y-th-p), (x+tw+p, y+bl+p), (8,8,8), -1)
    cv2.rectangle(img, (x-p, y-th-p), (x+tw+p, y+bl+p), color, 1)
    cv2.putText(img, text, (x, y), font, scale, color, thick, cv2.LINE_AA)

def draw_bone(img, p1, p2, color, thick=2):
    cv2.line(img, tuple(np.array(p1,int)), tuple(np.array(p2,int)), color, thick, cv2.LINE_AA)

def draw_joint(img, pt, color, r=4):
    cv2.circle(img, tuple(np.array(pt,int)), r, color, -1, cv2.LINE_AA)
    cv2.circle(img, tuple(np.array(pt,int)), r+1, (255,255,255), 1, cv2.LINE_AA)

def draw_arc(img, vertex, p1, p2, color, radius=18):
    v  = np.array(vertex, int)
    v1 = np.array(p1, float) - v
    v2 = np.array(p2, float) - v
    a1 = np.degrees(np.arctan2(v1[1], v1[0]))
    a2 = np.degrees(np.arctan2(v2[1], v2[0]))
    if a1 > a2: a1, a2 = a2, a1
    if a2 - a1 > 180: a1, a2 = a2, a1+360
    cv2.ellipse(img, tuple(v), (radius, radius), 0, a1, a2, color, 1, cv2.LINE_AA)

def offset_from(vertex, ref, dist=45):
    v = np.array(vertex, float); r = np.array(ref, float)
    d = v - r; nm = np.linalg.norm(d)
    if nm < 1: return v + np.array([dist, 0])
    return v + (d/nm)*dist

def draw_skeleton(image, lmd, scores, reba_final):
    img = image.copy()
    h, w = img.shape[:2]
    sp = get_scale_params(w, h)

    ear      = lmd['ear'];      nose     = lmd['nose']
    shoulder = lmd['shoulder']; elbow    = lmd['elbow']
    wrist    = lmd['wrist'];    hip      = lmd['hip']
    knee     = lmd['knee'];     ankle    = lmd['ankle']
    neck_s, trunk_s, leg_s, ua_s, la_s, w_s = scores
    main_c = risk_color_bgr(reba_final)

    bt = sp['bone']
    draw_bone(img, ear,      shoulder, seg_color(neck_s),  bt)
    draw_bone(img, shoulder, hip,      seg_color(trunk_s), bt+1)
    draw_bone(img, shoulder, elbow,    seg_color(ua_s),    bt)
    draw_bone(img, elbow,    wrist,    seg_color(la_s),    bt)
    draw_bone(img, hip,      knee,     seg_color(leg_s),   bt+1)
    draw_bone(img, knee,     ankle,    seg_color(leg_s),   bt+1)
    draw_bone(img, nose,     ear,      (160,160,160),      max(1, bt-1))

    head_r = max(8, int(abs(nose[1]-ear[1]) * sp['head']))
    cv2.circle(img, tuple(np.array(nose,int)), head_r, (200,200,200), 1, cv2.LINE_AA)

    jr = sp['joint']
    for pt, c in [(ear, seg_color(neck_s)), (shoulder, (220,220,100)),
                  (elbow, seg_color(ua_s)), (wrist, seg_color(la_s)),
                  (hip, (220,220,100)), (knee, seg_color(leg_s)), (ankle, seg_color(leg_s))]:
        draw_joint(img, pt, c, jr)

    ar = sp['arc']
    draw_arc(img, shoulder, ear,      hip,   seg_color(neck_s),  ar)
    draw_arc(img, hip,      shoulder, knee,  seg_color(trunk_s), int(ar*1.2))
    draw_arc(img, shoulder, hip,      elbow, seg_color(ua_s),    ar)
    draw_arc(img, elbow,    shoulder, wrist, seg_color(la_s),    int(ar*0.9))
    draw_arc(img, knee,     hip,      ankle, seg_color(leg_s),   ar)

    sc  = sp['font']; off = sp['offset']
    na  = lmd['neck_ang'];  ta   = lmd['trunk_ang']
    uaa = lmd['ua_ang'];    laa  = lmd['la_ang']
    wd  = lmd['wrist_dev']; ka   = lmd['knee_ang']

    def place(vertex, ref, text, color):
        pos = offset_from(vertex, ref, off)
        pos[0] = max(5, min(w-130, pos[0]))
        pos[1] = max(14, min(h-5,  pos[1]))
        draw_label(img, text, pos, color, sc)

    place(shoulder, ear,    f"Leher:{na:.1f}d[S{neck_s}]",  seg_color(neck_s))
    place(hip,      knee,   f"Tubuh:{ta:.1f}d[S{trunk_s}]", seg_color(trunk_s))
    place(elbow,    wrist,  f"L.Atas:{uaa:.1f}d[S{ua_s}]",  seg_color(ua_s))
    place(wrist,    elbow,  f"L.Bwh:{laa:.1f}d[S{la_s}]",   seg_color(la_s))
    place(knee,     ankle,  f"Lutut:{ka:.1f}d[S{leg_s}]",   seg_color(leg_s))
    place(wrist,    shoulder,f"Prglgn:{wd:.1f}d[S{w_s}]",   seg_color(w_s))

    kat, _, _ = risk_cat(reba_final)
    bfont  = sp['banner']
    banner = f"  REBA: {reba_final}  |  {kat}  "
    (bw2, bh), _ = cv2.getTextSize(banner, cv2.FONT_HERSHEY_SIMPLEX, bfont, 1)
    bx = (w - bw2) // 2
    pad_v = max(3, int(6*bfont))
    cv2.rectangle(img, (bx-8, 3), (bx+bw2+8, bh+pad_v*2), (12,12,12), -1)
    cv2.rectangle(img, (bx-8, 3), (bx+bw2+8, bh+pad_v*2), main_c, 1)
    cv2.putText(img, banner, (bx, bh+pad_v), cv2.FONT_HERSHEY_SIMPLEX, bfont, main_c, 1, cv2.LINE_AA)

    lfont  = sp['legend']
    legend = [((50,200,50),"S1"),((0,200,200),"S2"),((0,165,255),"S3"),((0,80,220),"S4+")]
    lx = 6; ly = h - 6
    for bgr, lbl in reversed(legend):
        (tw2, th2), _ = cv2.getTextSize(f" {lbl}", cv2.FONT_HERSHEY_SIMPLEX, lfont, 1)
        cv2.rectangle(img, (lx-2, ly-th2-3), (lx+tw2+14, ly+3), (12,12,12), -1)
        cv2.circle(img, (lx+4, ly-th2//2), 3, bgr, -1)
        cv2.putText(img, f"  {lbl}", (lx, ly), cv2.FONT_HERSHEY_SIMPLEX, lfont, bgr, 1, cv2.LINE_AA)
        ly -= th2 + 7
    return img

# =============================================
# ANALISIS POSE
# =============================================
def analyze_pose(image_bgr, beban, aktivitas, activity_score):
    """Jalankan REBA scoring. Return (annotated_bgr, result_dict) atau (None, None)."""
    results = pose_detector.process(cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB))
    if not results.pose_landmarks:
        return None, None

    h, w, _ = image_bgr.shape
    lm  = results.pose_landmarks.landmark
    LP  = mp_pose.PoseLandmark

    def gp(idx): return [lm[idx].x*w, lm[idx].y*h]

    ear      = gp(LP.LEFT_EAR.value);      nose     = gp(LP.NOSE.value)
    shoulder = gp(LP.LEFT_SHOULDER.value); elbow    = gp(LP.LEFT_ELBOW.value)
    wrist    = gp(LP.LEFT_WRIST.value);    hip      = gp(LP.LEFT_HIP.value)
    knee     = gp(LP.LEFT_KNEE.value);     ankle    = gp(LP.LEFT_ANKLE.value)

    ta  = trunk_flexion(shoulder, hip)
    na  = neck_flexion(ear, shoulder, hip)
    uaa = upper_arm_angle(shoulder, elbow, hip)
    laa = calc_angle(shoulder, elbow, wrist)
    ka  = calc_angle(hip, knee, ankle)
    wr  = calc_angle(elbow, wrist, shoulder)
    wd  = round(abs(180 - wr), 1)

    ns = score_neck(na);   ts = score_trunk(ta);  ls = score_legs(ka)
    us = score_ua(uaa);   las = score_la(laa);    ws = score_wrist(wd)

    tA = tbl_a(ts, ns, ls); tB = tbl_b(us, las, ws)
    fs = force_score(beban); cs = 1
    sA = tA + fs;  sB = tB + cs
    sC = tbl_c(sA, sB)
    final = max(1, min(15, sC + activity_score))
    kat, warna, tindakan = risk_cat(final)

    lmd = {
        'ear':ear,'nose':nose,'shoulder':shoulder,'elbow':elbow,
        'wrist':wrist,'hip':hip,'knee':knee,'ankle':ankle,
        'neck_ang':na,'trunk_ang':ta,'ua_ang':uaa,'la_ang':laa,
        'wrist_dev':wd,'knee_ang':ka,
    }
    annotated = draw_skeleton(image_bgr, lmd, (ns, ts, ls, us, las, ws), final)

    result = {
        "Timestamp"                : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Beban (kg)"               : beban,
        "Aktivitas"                : aktivitas,
        "Sudut Leher (deg)"        : na,
        "Sudut Batang Tubuh (deg)" : ta,
        "Sudut Lengan Atas (deg)"  : uaa,
        "Sudut Lengan Bawah (deg)" : laa,
        "Deviasi Pergelangan (deg)": wd,
        "Sudut Lutut (deg)"        : ka,
        "Skor Leher"               : ns,
        "Skor Batang Tubuh"        : ts,
        "Skor Kaki"                : ls,
        "Skor Lengan Atas"         : us,
        "Skor Lengan Bawah"        : las,
        "Skor Pergelangan"         : ws,
        "Table A"                  : tA,
        "Table B"                  : tB,
        "Force/Load Score"         : fs,
        "Score A"                  : sA,
        "Score B"                  : sB,
        "Score C"                  : sC,
        "Activity Score"           : activity_score,
        "Skor REBA Final"          : final,
        "Kategori Risiko"          : kat,
        "Tindakan"                 : tindakan,
    }
    return annotated, result

# =============================================
# EXPORT EXCEL
# =============================================
def build_excel(result: dict, nama_aktivitas: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data_export = dict(result)
    data_export["Nama Aktivitas (Manual)"] = nama_aktivitas
    ordered = ["Timestamp","Nama Aktivitas (Manual)","Beban (kg)","Aktivitas"] + \
              [k for k in data_export if k not in
               ["Timestamp","Nama Aktivitas (Manual)","Beban (kg)","Aktivitas"]]
    data_export = {k: data_export[k] for k in ordered if k in data_export}

    df = pd.DataFrame([data_export])
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as wr:
        df.to_excel(wr, sheet_name="Hasil REBA", index=False, startrow=2)
        ws = wr.sheets["Hasil REBA"]

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Laporan REBA — {nama_aktivitas} — {datetime.now().strftime('%d %B %Y')}"
        ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor="1A3A6A")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for cell in ws[3]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor="2E4A7A")
            cell.alignment = Alignment(horizontal="center")

        reba_col = None
        for i, cell in enumerate(ws[3], 1):
            if cell.value == "Skor REBA Final": reba_col = i; break
        if reba_col:
            score = result["Skor REBA Final"]
            cmap = {(1,1):"27AE60",(2,3):"2ECC71",(4,7):"F39C12",(8,10):"E67E22",(11,15):"E74C3C"}
            cc = "FFFFFF"
            for (lo, hi), hx in cmap.items():
                if lo <= score <= hi: cc = hx; break
            dc = ws.cell(row=4, column=reba_col)
            dc.fill = PatternFill("solid", fgColor=cc)
            dc.font = Font(bold=True, color="000000")

        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )
        for row in ws.iter_rows(min_row=3, max_row=4):
            for cell in row: cell.border = thin

        for col in ws.columns:
            ml = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(ml+3, 12)

    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
if "result"     not in st.session_state: st.session_state.result     = None
if "annotated"  not in st.session_state: st.session_state.annotated  = None
if "analyzed"   not in st.session_state: st.session_state.analyzed   = False

# =============================================
# SIDEBAR
# =============================================
with st.sidebar:
    st.markdown('<div class="reba-title">REBA</div>', unsafe_allow_html=True)
    st.markdown('<div class="reba-subtitle">Ergonomics Analyzer 2026</div>', unsafe_allow_html=True)
    st.divider()

    st.markdown("#### 🖼️ Upload Foto")
    uploaded = st.file_uploader(
        "Pilih gambar (JPG / PNG / BMP)",
        type=["jpg","jpeg","png","bmp"],
        label_visibility="collapsed",
    )

    st.divider()

    # ── Parameter REBA ───────────────────────────────────────────────────────
    st.markdown("#### ⚖️ Beban yang Diangkat")
    beban = st.number_input("Berat (kg)", min_value=0.0, max_value=200.0,
                             value=0.0, step=0.5, format="%.1f")
    fs_val = force_score(beban)
    force_labels = {0: "🟢 Ringan (< 5 kg) — Skor 0",
                    1: "🟡 Sedang (5–10 kg) — Skor 1",
                    2: "🔴 Berat (> 10 kg) — Skor 2"}
    st.caption(force_labels[fs_val])

    st.divider()
    st.markdown("#### 🏭 Nama Aktivitas")
    AKTIVITAS_LIST = [
        "-- Pilih Aktivitas --",
        "Pengangkatan Manual",
        "Menurunkan Beban",
        "Mendorong / Menarik",
        "Perakitan (Assembly)",
        "Pengelasan",
        "Pengepakan / Packaging",
        "Inspeksi Visual",
        "Pengoperasian Mesin",
        "Pekerjaan Kantor / Duduk",
        "Lainnya (isi manual)...",
    ]
    aktivitas_sel = st.selectbox("Aktivitas", AKTIVITAS_LIST, label_visibility="collapsed")
    aktivitas_manual = ""
    if aktivitas_sel == "Lainnya (isi manual)...":
        aktivitas_manual = st.text_input("Tulis nama aktivitas...")
    aktivitas = aktivitas_manual.strip() if aktivitas_sel == "Lainnya (isi manual)..." \
                else (aktivitas_sel if aktivitas_sel != "-- Pilih Aktivitas --" else "Tidak disebutkan")

    st.divider()
    st.markdown("#### 🔢 Activity Score  *(Step 13 REBA)*")
    st.caption("Setiap kondisi yang berlaku menambah +1 ke Score C.")
    act1 = st.checkbox("🧍 1 atau lebih bagian tubuh statis\n*(ditahan > 1 menit)*")
    act2 = st.checkbox("🔄 Gerakan berulang kecil\n*(> 4×/menit, bukan berjalan)*")
    act3 = st.checkbox("⚡ Postur berubah cepat / tidak stabil")
    activity_score = sum([act1, act2, act3])
    st.info(f"Activity Score Total: **+{activity_score}**")

    st.divider()
    analyze_btn = st.button("🔍  Analisis REBA", type="primary",
                             use_container_width=True,
                             disabled=(uploaded is None))

    # Info alur kerja
    st.divider()
    st.markdown("""
**Alur Kerja:**
1. Upload foto postur kerja
2. Isi beban & nama aktivitas
3. Centang activity score
4. Klik **Analisis REBA**
5. Download hasil Excel
""")
    st.caption("Hignett & McAtamney\nApplied Ergonomics 31 (2000)")

# =============================================
# MAIN AREA
# =============================================
col_img, col_result = st.columns([3, 2], gap="medium")

with col_img:
    st.markdown("### 🦴 Visualisasi Skeleton & Label Sudut")
    if uploaded is not None and not st.session_state.analyzed:
        pil_img = Image.open(uploaded).convert("RGB")
        st.image(pil_img, caption="Pratinjau gambar yang diupload", use_container_width=True)

    if st.session_state.analyzed and st.session_state.annotated is not None:
        annotated_rgb = cv2.cvtColor(st.session_state.annotated, cv2.COLOR_BGR2RGB)
        st.image(annotated_rgb, caption="Hasil anotasi skeleton REBA", use_container_width=True)
    elif not uploaded:
        st.info("⬅️ Upload foto postur kerja terlebih dahulu melalui panel kiri, "
                "kemudian klik **Analisis REBA**.")

with col_result:
    st.markdown("### 📊 Hasil Analisis REBA")
    if st.session_state.analyzed and st.session_state.result:
        r = st.session_state.result
        reba = r["Skor REBA Final"]
        kat, warna, tindakan = risk_cat(reba)

        # Kartu skor utama
        st.markdown(
            f'<div class="score-card" style="background:{warna}22; border:2px solid {warna};">'
            f'<div style="font-size:2.5rem;font-weight:800;color:{warna};">{reba}</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:{warna};">Risiko {kat}</div>'
            f'<div style="font-size:.85rem;color:#CCCCDD;">{tindakan}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # Sudut terukur
        st.markdown("**📐 Sudut Terukur**")
        sudut_data = {
            "Segmen": ["Leher","Batang Tubuh","Lengan Atas","Lengan Bawah","Pergelangan","Lutut"],
            "Sudut (°)": [
                r["Sudut Leher (deg)"], r["Sudut Batang Tubuh (deg)"],
                r["Sudut Lengan Atas (deg)"], r["Sudut Lengan Bawah (deg)"],
                r["Deviasi Pergelangan (deg)"], r["Sudut Lutut (deg)"],
            ],
            "Skor": [
                r["Skor Leher"], r["Skor Batang Tubuh"],
                r["Skor Lengan Atas"], r["Skor Lengan Bawah"],
                r["Skor Pergelangan"], r["Skor Kaki"],
            ],
        }
        st.dataframe(pd.DataFrame(sudut_data), hide_index=True, use_container_width=True)

        # Alur perhitungan
        st.markdown("**🔢 Alur Perhitungan**")
        alur_data = {
            "Parameter": ["Table A","Force Score","Score A",
                          "Table B","Coupling","Score B",
                          "Score C","Activity Score","REBA FINAL"],
            "Nilai": [
                r["Table A"], r["Force/Load Score"], r["Score A"],
                r["Table B"], "1 (Fair)",            r["Score B"],
                r["Score C"], f"+{r['Activity Score']}", r["Skor REBA Final"],
            ],
        }
        st.dataframe(pd.DataFrame(alur_data), hide_index=True, use_container_width=True)

        # Info
        st.markdown("**📋 Info Sesi**")
        st.caption(f"🕐 {r['Timestamp']}  |  ⚖️ {r['Beban (kg)']} kg  |  🏭 {r['Aktivitas']}")

        # Export Excel
        st.divider()
        st.markdown("**📤 Export ke Excel**")
        nama_export = st.text_input(
            "Nama Aktivitas untuk Laporan",
            value=r.get("Aktivitas", ""),
            placeholder="Contoh: Pengangkatan Barang Gudang A",
        )
        if nama_export.strip():
            excel_bytes = build_excel(r, nama_export.strip())
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            st.download_button(
                label="⬇️  Download Excel",
                data=excel_bytes,
                file_name=f"REBA_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    else:
        st.info("Hasil analisis akan muncul di sini setelah foto diproses.")

# =============================================
# PROSES ANALISIS (dipanggil saat tombol diklik)
# =============================================
if analyze_btn and uploaded is not None:
    pil_img = Image.open(uploaded).convert("RGB")
    img_bgr = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)

    with st.spinner("Mendeteksi pose & menghitung skor REBA..."):
        annotated, result = analyze_pose(img_bgr, beban, aktivitas, activity_score)

    if result is None:
        st.error(
            "⚠️ **Pose tidak terdeteksi** dalam gambar yang diupload.\n\n"
            "Pastikan:\n"
            "- Seluruh tubuh (kepala hingga kaki) terlihat jelas\n"
            "- Pencahayaan cukup\n"
            "- Tidak ada objek yang menghalangi pose"
        )
        st.session_state.analyzed  = False
        st.session_state.result    = None
        st.session_state.annotated = None
    else:
        st.session_state.result    = result
        st.session_state.annotated = annotated
        st.session_state.analyzed  = True
        st.rerun()
